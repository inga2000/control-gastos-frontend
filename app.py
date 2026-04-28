import streamlit as st
import pandas as pd
import requests
from datetime import date
import calendar
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

# =====================
# Configuración
# =====================
st.set_page_config(page_title="Control de gastos", layout="wide")
st.title("💰 Control de gastos")

API_URL = "https://control-gastos-backend-4s5z.onrender.com"

# =====================
# Categorías UX
# =====================
CATEGORIAS = {
    "🚗 Transporte": [
        "Combustible", "Transporte público", "Estacionamiento",
        "Peajes", "Mantenimiento auto", "Seguro auto"
    ],
    "🏠 Hogar": [
        "Alquiler", "Luz", "Gas", "Agua",
        "Internet", "Expensas", "Impuestos"
    ],
    "🍕 Comida": [
        "Supermercado", "Comidas fuera", "Delivery"
    ],
    "💊 Salud": [
        "Farmacia", "Prepaga", "Consultas médicas"
    ],
    "📱 Servicios": [
        "Celular", "Streaming", "Otros servicios"
    ],
    "🛍️ Gustos": [
        "Ropa", "Salidas", "Viajes", "Regalos", "Otros"
    ],
    "💰 Ingresos": [
        "Sueldo", "Freelance", "Extras", "Otros ingresos"
    ]
}

# =====================
# Backend helper
# =====================
def cargar_datos():
    r = requests.get(f"{API_URL}/movimientos")
    if r.status_code == 200:
        df = pd.DataFrame(r.json())
        if not df.empty:
            df.rename(columns={
                "fecha": "Fecha",
                "monto": "Monto",
                "tipo": "Tipo",
                "categoria": "Categoría",
                "descripcion": "Descripción"
            }, inplace=True)
            df["Fecha"] = pd.to_datetime(df["Fecha"])
        return df
    return pd.DataFrame()

datos = cargar_datos()

# =====================
# Tabs
# =====================
tab_mes, tab_hist = st.tabs(["📅 Mes actual", "📊 Histórico"])

# ==========================================================
# TAB 1 — MES ACTUAL
# ==========================================================
with tab_mes:

    # -------- Alta --------
    st.header("➕ Agregar movimiento")

    c1, c2, c3 = st.columns(3)
    with c1:
        fecha = st.date_input("📅 Fecha", value=date.today(), format="DD/MM/YYYY")
    with c2:
        monto = st.number_input("💵 Monto", min_value=0.0, step=100.0)
    with c3:
        tipo = st.selectbox("Tipo", ["Gasto", "Ingreso"])

    c4, c5 = st.columns(2)
    with c4:
        grupo = st.selectbox("Grupo", list(CATEGORIAS.keys()))
    with c5:
        subcategoria = st.selectbox("Categoría", CATEGORIAS[grupo])

    descripcion = st.text_input("📝 Descripción")

    if st.button("✅ Guardar movimiento"):
        payload = {
            "fecha": str(fecha),
            "monto": float(monto),
            "tipo": tipo,
            "categoria": f"{grupo} › {subcategoria}",
            "descripcion": descripcion
        }
        requests.post(f"{API_URL}/movimientos", json=payload)
        st.rerun()

    # -------- Calendario --------
    if not datos.empty:
        st.header("📅 Calendario mensual")

        mes = st.selectbox("Mes", range(1, 13), index=date.today().month - 1)
        anio = st.selectbox("Año", sorted(datos["Fecha"].dt.year.unique(), reverse=True))

        datos_mes = datos[
            (datos["Fecha"].dt.month == mes) &
            (datos["Fecha"].dt.year == anio)
        ].copy()

        total_ing = datos_mes[datos_mes["Tipo"] == "Ingreso"]["Monto"].sum()
        total_gas = datos_mes[datos_mes["Tipo"] == "Gasto"]["Monto"].sum()

        c1, c2, c3 = st.columns(3)
        c1.metric("🟢 Ingresos", f"${total_ing:,.0f}")
        c2.metric("🔴 Gastos", f"${total_gas:,.0f}")
        c3.metric("⚖️ Balance", f"${total_ing - total_gas:,.0f}")

        resumen = datos_mes.copy()
        resumen["Ingreso"] = resumen.apply(lambda x: x["Monto"] if x["Tipo"] == "Ingreso" else 0, axis=1)
        resumen["Gasto"] = resumen.apply(lambda x: x["Monto"] if x["Tipo"] == "Gasto" else 0, axis=1)
        diario = resumen.groupby(resumen["Fecha"].dt.day)[["Ingreso", "Gasto"]].sum()

        cal = calendar.monthcalendar(anio, mes)
        filas = []

        for semana in cal:
            fila = []
            for dia in semana:
                if dia == 0:
                    fila.append("")
                elif dia in diario.index:
                    fila.append(
                        f"{dia}\n"
                        f"- {diario.loc[dia,'Gasto']:.0f}\n"
                        f"+ {diario.loc[dia,'Ingreso']:.0f}"
                    )
                else:
                    fila.append(str(dia))
            filas.append(fila)

        st.dataframe(
            pd.DataFrame(filas, columns=["Lun","Mar","Mié","Jue","Vie","Sáb","Dom"]),
            use_container_width=True
        )

        # -------- Detalle / Editar / Borrar --------
        st.header("📋 Detalle por día")

        if not diario.empty:
            dia_sel = st.selectbox("Día", sorted(diario.index))
            detalle = datos_mes[datos_mes["Fecha"].dt.day == dia_sel]
            detalle_vis = detalle.copy()
            detalle_vis["Fecha"] = detalle_vis["Fecha"].dt.strftime("%d/%m/%Y")
            st.dataframe(detalle_vis, use_container_width=True)

            mov_ids = detalle.index.tolist()
            mov_idx = st.selectbox(
                "Movimiento",
                mov_ids,
                format_func=lambda i: f"{detalle.loc[i,'Tipo']} ${detalle.loc[i,'Monto']}"
            )
            mov = detalle.loc[mov_idx]

            ef = st.date_input("Fecha", mov["Fecha"].date(), format="DD/MM/YYYY")
            em = st.number_input("Monto", value=float(mov["Monto"]))
            et = st.selectbox("Tipo", ["Gasto","Ingreso"], index=0 if mov["Tipo"]=="Gasto" else 1,key=f"edit_tipo_{mov['id']}")
            ec = st.text_input("Categoría", mov["Categoría"])
            ed = st.text_input("Descripción", mov["Descripción"])

            c1, c2 = st.columns(2)
            with c1:
                if st.button("💾 Guardar cambios"):
                    requests.put(f"{API_URL}/movimientos/{mov['id']}", json={
                        "fecha": str(ef),
                        "monto": em,
                        "tipo": et,
                        "categoria": ec,
                        "descripcion": ed
                    })
                    st.rerun()

            with c2:
                if st.button("🗑️ Borrar movimiento"):
                    requests.delete(f"{API_URL}/movimientos/{mov['id']}")
                    st.rerun()

        # -------- Excel (CALENDARIO COMPLETO) --------
        st.header("📥 Descargar calendario mensual en Excel")

        buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Calendario"

        fill_ing = PatternFill("solid", fgColor="C6EFCE")
        fill_gas = PatternFill("solid", fgColor="F8CBAD")
        fill_head = PatternFill("solid", fgColor="D9D9D9")

        headers = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]
        for col,h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = fill_head
            cell.font = Font(bold=True)
            ws.column_dimensions[get_column_letter(col)].width = 22

        ingresos = datos_mes[datos_mes["Tipo"]=="Ingreso"].groupby(datos_mes["Fecha"].dt.day)["Monto"].sum()
        gastos = datos_mes[datos_mes["Tipo"]=="Gasto"].groupby(datos_mes["Fecha"].dt.day)["Monto"].sum()

        fila_excel = 2
        for semana in cal:
            for col, dia in enumerate(semana, start=1):
                if dia == 0:
                    continue
                c = ws.cell(row=fila_excel, column=col)
                texto = f"{dia}"
                if dia in ingresos:
                    texto += f"\n+ {ingresos[dia]:.0f}"
                    c.fill = fill_ing
                if dia in gastos:
                    texto += f"\n- {gastos[dia]:.0f}"
                    c.fill = fill_gas
                c.value = texto
                c.alignment = Alignment(wrap_text=True, vertical="top")
            fila_excel += 1

        wb.save(buffer)

        st.download_button(
            "📅 Descargar Excel",
            buffer.getvalue(),
            f"calendario_{anio}_{mes:02d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# ==========================================================
# TAB 2 — HISTÓRICO
# ==========================================================
with tab_hist:
    st.header("📊 Histórico")

    if datos.empty:
        st.info("No hay datos todavía")
    else:
        col1, col2 = st.columns(2)
        with col1:
            anio_hist = st.selectbox("Año", sorted(datos["Fecha"].dt.year.unique(), reverse=True), key="hist_anio")
        with col2:
            meses_disp = sorted(datos[datos["Fecha"].dt.year == anio_hist]["Fecha"].dt.month.unique())
            mes_hist = st.selectbox("Mes", ["Todos"] + meses_disp, key="hist_mes")

        datos_hist = datos[datos["Fecha"].dt.year == anio_hist]
        if mes_hist != "Todos":
            datos_hist = datos_hist[datos_hist["Fecha"].dt.month == mes_hist]

        datos_vis = datos_hist.copy()
        datos_vis["Fecha"] = datos_vis["Fecha"].dt.strftime("%d/%m/%Y")
        st.dataframe(datos_vis.sort_values("Fecha"), use_container_width=True)

        ti = datos_hist[datos_hist["Tipo"]=="Ingreso"]["Monto"].sum()
        tg = datos_hist[datos_hist["Tipo"]=="Gasto"]["Monto"].sum()

        c1, c2, c3 = st.columns(3)
        c1.metric("🟢 Ingresos", f"${ti:,.0f}")
        c2.metric("🔴 Gastos", f"${tg:,.0f}")
        c3.metric("⚖️ Balance", f"${ti-tg:,.0f}")